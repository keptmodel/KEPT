import os

from pandas import DataFrame
from sklearn.metrics import precision_recall_curve, PrecisionRecallDisplay
import matplotlib.pyplot as plt
import openpyxl


class metrics:
    def __init__(self, data_frame: DataFrame, output_dir=None):
        """
        Evaluate the performance given datafrome with column "s_id", "t_id" "pred" and "label"
        :param data_frame:
        """
        self.data_frame = data_frame
        self.output_dir = output_dir
        self.s_ids, self.t_ids = data_frame['s_id'], data_frame['t_id']
        self.pred, self.label = data_frame['pred'], data_frame['label']
        self.group_sort = None

    def f1_score(self, precision, recall):
        return 2 * (precision * recall) / (precision + recall) if precision + recall > 0 else 0

    def f2_score(self, precision, recall):
        return 5 * precision * recall / (4 * precision + recall) if precision + recall > 0 else 0

    def f1_details(self, threshold):
        "Return ture positive (tp), fp, tn,fn "
        f_name = "f1_details"
        tp, fp, tn, fn = 0, 0, 0, 0
        for p, l in zip(self.pred, self.label):
            if p > threshold:
                p = 1
            else:
                p = 0
            if p == l:
                if l == 1:
                    tp += 1
                else:
                    tn += 1
            else:
                if l == 1:
                    fp += 1
                else:
                    fn += 1
        return {"tp": tp, "fp": fp, "tn": tn, "fn": fn}

    def precision_recall_curve(self, fig_name):
        precision, recall, thresholds = precision_recall_curve(self.label, self.pred)
        max_f1 = 0
        max_f2 = 0
        max_threshold = 0
        for p, r, tr in zip(precision, recall, thresholds):
            f1 = self.f1_score(p, r)
            f2 = self.f2_score(p, r)
            if f1 >= max_f1:
                max_f1 = f1
                max_threshold = tr
            if f2 >= max_f2:
                max_f2 = f2
        viz = PrecisionRecallDisplay(
            precision=precision, recall=recall)
        viz.plot()
        if os.path.isdir(self.output_dir):
            fig_path = os.path.join(self.output_dir, fig_name)
            plt.savefig(fig_path)
            plt.close()
        detail = self.f1_details(max_threshold)
        return round(max_f1, 3), round(max_f2, 3), detail, max_threshold

    def precision_at_K(self, k=1):
        if self.group_sort is None:
            self.group_sort = self.data_frame.groupby(["s_id"]).apply(
                lambda x: x.sort_values(["pred"], ascending=False)).reset_index(drop=True)
        group_tops = self.group_sort.groupby('s_id')
        cnt = 0
        hits = 0
        for s_id, group in group_tops:
            hit = 0
            for index, row in group.head(k).iterrows():
                hit += 1 if row['label'] == 1 else 0
            hits += float(hit) / k
            cnt += 1
        return round(hits / cnt if cnt > 0 else 0, 3)

    def recall_at_K(self, k=1):
        if self.group_sort is None:
            self.group_sort = self.data_frame.groupby(["s_id"]).apply(
                lambda x: x.sort_values(["pred"], ascending=False)).reset_index(drop=True)
        group_tops = self.group_sort.groupby('s_id')
        cnt = 0
        hits = 0
        for s_id, group in group_tops:
            sum1 = group['label'].sum()
            hit = 0
            for index, row in group.head(k).iterrows():
                hit += 1 if row['label'] == 1 else 0
            hits += float(hit) / sum1
            cnt += 1
        return round(hits / cnt if cnt > 0 else 0, 3)

    def top_at_K(self, k=1):
        if self.group_sort is None:
            self.group_sort = self.data_frame.groupby(["s_id"]).apply(
                lambda x: x.sort_values(["pred"], ascending=False)).reset_index(drop=True)
        group_tops = self.group_sort.groupby('s_id')
        cnt = 0
        hits = 0
        for s_id, group in group_tops:
            for index, row in group.head(k).iterrows():
                if row['label'] == 1:
                    hits += 1
                    break
            cnt += 1
        return round(hits / cnt if cnt > 0 else 0, 3)

    def MAP_at_K(self):
        if self.group_sort is None:
            self.group_sort = self.data_frame.groupby(["s_id"]).apply(
                lambda x: x.sort_values(["pred"], ascending=False)).reset_index(drop=True)
        group_tops = self.group_sort.groupby('s_id')
        ap_sum = 0
        for s_id, group in group_tops:
            group_hits = 0
            sum1=group['label'].sum()
            ap = 0
            for i, (index, row) in enumerate(group.iterrows()):
                if row['label'] == 1:
                    group_hits += 1
                    ap += (group_hits) / (i + 1) / sum1
                    if group_hits>=sum1:
                        break
            ap_sum += ap
        map = ap_sum / len(group_tops) if len(group_tops) > 0 else 0
        return round(map, 3)

    def MRR(self):
        if self.group_sort is None:
            self.group_sort = self.data_frame.groupby(["s_id"]).apply(
                lambda x: x.sort_values(["pred"], ascending=False)).reset_index(drop=True)
        group_tops = self.group_sort.groupby('s_id')
        mrr_sum = 0
        for s_id, group in group_tops:
            rank = 0
            for i, (index, row) in enumerate(group.iterrows()):
                rank += 1
                if row['label'] == 1:
                    mrr_sum += 1.0 / rank
                    break
        return round(mrr_sum / len(group_tops), 3)

    def get_all_metrices(self):
        pk1 = self.precision_at_K(1)
        pk3 = self.precision_at_K(3)
        pk5 = self.precision_at_K(5)
        pk10 = self.precision_at_K(10)
        rk1 = self.recall_at_K(1)
        rk3 = self.recall_at_K(3)
        rk5 = self.recall_at_K(5)
        rk10 = self.recall_at_K(10)
        top1 = self.top_at_K(1)
        top3 = self.top_at_K(3)
        top5 = self.top_at_K(5)
        top10 = self.top_at_K(10)

        map = self.MAP_at_K()
        mrr = self.MRR()
        return {
            'pk1': pk1,
            'pk3': pk3,
            'pk5': pk5,
            'pk10': pk10,
            'rk1': rk1,
            'rk3': rk3,
            'rk5': rk5,
            'rk10': rk10,
            'top1': top1,
            'top3': top3,
            'top5': top5,
            'top10': top10,
            'map': map,
            'mrr': mrr,
        }

    def write_summary(self, exe_time):
        summary_path = os.path.join(self.output_dir, "summary.txt")
        res = self.get_all_metrices()
        pk10, pk5, pk3, pk1 = res['pk10'], res['pk5'], res['pk3'], res['pk1']
        rk10, rk5, rk3, rk1 = res['rk10'], res['rk5'], res['rk3'], res['rk1']
        top10, top5, top3, top1 = res['top10'], res['top5'], res['top3'], res['top1']
        map, mrr = res['map'], res['mrr']
        summary = "\npk10={}, pk5={},pk3={},pk1 = {}, rk10={},rk5={},rk3={},rk1={},top10={},top5={},top3={},top1={}, MAP={}, MRR={}, exe_time={}\n".format(
            pk10,
            pk5,
            pk3,
            pk1,
            rk10,
            rk5,
            rk3,
            rk1,
            top10,
            top5,
            top3,
            top1,
            map,
            mrr,
            exe_time,
        )
        result_list = [pk10, pk5, pk3, pk1, rk10, rk5, rk3, rk1, top10, top5, top3, top1, map, mrr]
        #workbook = openpyxl.load_workbook('result.xlsx')
        #worksheet = workbook.active
        #worksheet.append(result_list)
        #workbook.save('result.xlsx')
        #workbook.close()
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        for i, result in enumerate(result_list, start=1):
            worksheet.cell(row=i, column=1, value=result)
        workbook.save(os.path.join(self.output_dir,'result.xlsx'))
        workbook.close()
        with open(summary_path, 'w') as fout:
            fout.write(summary)
        print(summary)
